import openpyxl

class XlsUtils:
  # _instance = None
  _sheet = None

  # def __new__(cls):
  #   if cls._instance is None:
  #     cls._instance = object.__new__(cls)
  #   return cls._instance

  def open_sheet(self, file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    self._sheet = workbook[sheet_name]

  def get_cell(self, row=None, col=None, pos=None):
    if pos:
      cell = self._sheet.cell(row=int(pos[0]), column=int(pos[1]))
      return self._sheet.cell(row=int(pos[0]), column=int(pos[1])).value
    if row and col:
      return self._sheet.cell(row=int(row), column=int(col)).value
    return None